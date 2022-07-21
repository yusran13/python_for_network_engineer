from openpyxl import Workbook, load_workbook

#Read data from xlsx file
# grab the active worksheet
wb = load_workbook(filename = 'All Devices Jul 21, 2022 10.41.30 AM.xlsx')
active_sheet = wb['Sheet1']
print(f"Maximum column = {active_sheet.max_column}")

print(active_sheet['A1'].value)
print(active_sheet['B1'].value)
print(active_sheet['C1'].value)

print(active_sheet.cell(row=1, column=1).value)
print(active_sheet.cell(row=1, column=2).value)
print(active_sheet.cell(row=1, column=3).value)


#Write data to xlsx file
# new_file = Workbook()
# active_sheet = new_file.active
# active_sheet.cell(row=1, column=1).value = 'No'
# active_sheet.cell(row=1, column=2).value = 'Site ID'
# active_sheet.cell(row=1, column=3).value = 'Sitename'
# active_sheet.cell(row=1, column=4).value = 'Hostname'

# new_file.save('Test_new_file_from_python.xlsx')